import streamlit as st
import pandas as pd
import numpy as np
import json
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# ---------- 核心处理函数 ----------
def generate_portrait(json_files):
    # 1. 固定从仓库根目录读取映射表和模板（与 app.py 同级）
    mapping_path = 'data/try_3.0.xlsx'
    template_path = 'data/多人群画像_模板_3.0.xlsx'
    
    # 检查文件是否存在（方便调试）
    if not os.path.exists(mapping_path):
        raise FileNotFoundError(f"映射表文件 '{mapping_path}' 未找到，请确认已上传到仓库根目录")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件 '{template_path}' 未找到，请确认已上传到仓库根目录")
    
    # 读取映射表，只保留 'A' 列
    df_try = pd.read_excel(mapping_path)
    if 'A' not in df_try.columns:
        raise ValueError("映射表必须包含 'A' 列")
    df_try = df_try[['A']].copy()

    # 2. 遍历所有上传的 JSON 文件
    for uploaded_file in json_files:
        name = uploaded_file.name.replace('.txt', '').replace('-', '_')
        raw = pd.read_json(uploaded_file)
        raw_new = raw.explode('snapshotResponseRegions')

        # 展开 tagValueResults
        result_df = pd.DataFrame()
        for _, row in raw_new.iterrows():
            rows_to_append = row['snapshotResponseRegions']['tagValueResults']
            result_df = pd.concat([result_df, pd.DataFrame(rows_to_append)], ignore_index=True)

        # 提取三列
        result_df_final = pd.DataFrame({
            'tagName': result_df['tagEname'],
            'tagValueName': result_df['tagValueName'],
            'rate': result_df['rate']
        })

        # 中文映射
        replace_map = {
            'cosmetics_year_cnt_region': '彩妆年消费频次',
            'skin_year_amt_region': '护肤年消费金额',
            'daas_tag_pred_gender': '性别',
            'common_receive_city_level_180d': '城市等级',
            'pref_purchasing_power': '购买力',
            'skin_year_cnt_region': '护肤年消费频次',
            'cosmetics_year_amt_region': '彩妆年消费金额',
            'daas_tag_pred_age_level': '年龄',
            'dkx_strategy_crowd': '策略人群',
            'derive_pay_ord_amt_6m_015_range': '月均消费金额',
            'daas_tag_label_name': '礼遇场景人群',
            'daas_tag_gift_label_name': '礼遇人群',
            'daas_tag_resident_city_tb_userid': '预测城市',
            'common_receive_province_180d': '预测省份',
            'daas_tag_cosmetic_market': '美妆洗护市场',
            'pref_beauty_skin': '肤质',
            'daas_tag_purchase_driven_factor_tag': '美妆心智人群',
            'daas_tag_trade_hour_most': '预测购买时间',
            'pref_skincare': '护肤品功效需求',
            'daas_tag_kx_crowd_name': '十大策略人群',
            'pref_cosmetics': '彩妆功效需求',
            'daas_tag_mz_seven_crowd': '美妆七大功效人群'
        }
        result_df_final['tagName'] = result_df_final['tagName'].replace(replace_map)
        # 特殊处理性别、年龄
        result_df_final.loc[result_df_final['tagName'].str.contains('daas_tag_pred_gender'), 'tagName'] = '性别'
        result_df_final.loc[result_df_final['tagName'].str.contains('daas_tag_pred_age_level'), 'tagName'] = '年龄'

        # 构造 A、B 列
        df10 = result_df_final[['tagName', 'tagValueName', 'rate']].copy()
        df10.columns = ['A', 'B', 'C']
        df10['A'] = df10['A'] + df10['B']
        df10['B'] = df10['C']
        df10.drop('C', axis=1, inplace=True)

        # 计算人群量级
        ttl_df = result_df[result_df['tagEname'] == 'pref_purchasing_power']
        ttl = ttl_df['count'].sum() if not ttl_df.empty else 0
        ttl_row = pd.DataFrame([['人群量级', ttl]], columns=['A', 'B'])

        result_t = pd.concat([ttl_row, df10], ignore_index=True)
        result_t_deduplicated = result_t.groupby('A', as_index=False).first()

        # 左连接
        df_try = pd.merge(df_try, result_t_deduplicated, how='left', on='A')
        df_try[name] = df_try['B']
        df_try.drop(columns=['B'], inplace=True)

    # ---------- 3. 生成输出文件 ----------
    wb = openpyxl.load_workbook(template_path)
    
    if "1" in wb.sheetnames:
        wb.remove(wb["1"])
    
    ws = wb.create_sheet("1")
    
    for r in dataframe_to_rows(df_try, index=False, header=True):
        ws.append(r)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------- Streamlit UI ----------
st.set_page_config(page_title="人群画像生成器", layout="centered")
st.title("🧑‍🤝‍🧑 人群画像轻量生成器")

st.markdown("""
**使用说明：**
- 映射表（`try_3.0.xlsx`）和模板文件已内置在系统中，无需上传。
- 请上传需要处理的 JSON 文件（.txt 格式），可多选。
- 点击按钮后自动生成包含所有标签画像的新 Excel 文件。
""")

with st.form("upload_form"):
    json_files = st.file_uploader(
        "📁 上传 JSON 文件（可多选 .txt）", 
        type=["txt"],
        accept_multiple_files=True,
        help="选择你需要处理的所有 JSON 文本文件，可以按住 Ctrl 多选。"
    )
    submitted = st.form_submit_button("🚀 开始生成画像")

if submitted:
    if len(json_files) == 0:
        st.error("请至少上传一个 JSON 文件！")
    else:
        with st.spinner("正在处理，请稍候..."):
            try:
                output_bytes = generate_portrait(json_files)
                now = datetime.now()
                out_filename = now.strftime("%Y%m%d_%H%M%S") + ".xlsx"
                st.success("✅ 生成成功！点击下方按钮下载结果。")
                st.download_button(
                    label="📥 下载画像 Excel",
                    data=output_bytes,
                    file_name=out_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"运行出错：{e}")
                st.exception(e)
